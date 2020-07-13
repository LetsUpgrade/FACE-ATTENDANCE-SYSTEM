"""
@author: Meet Vansjaliya
"""

import face_recognition
import os
from face_recognition.face_recognition_cli import image_files_in_folder
import sys
import cv2

my_dir = os.fsencode('/home/sintu/Downloads/Project_pic_training/Ashutosh_Pandey/') # Folder where all your image files reside. Ensure it ends with '/
encoding_for_ashutosh = [] # Create an empty list for saving encoded files
for a in os.listdir(my_dir): # Loop over the folder to list individual files
    image = my_dir + a
    image = face_recognition.load_image_file(image) # Run your load command
    image_encoding = face_recognition.face_encodings(image) # Run your encoding command
    encoding_for_ashutosh.append(image_encoding[0]) # Append the results to encoding_for_file list
    
    
    
    
my_dir1 = os.fsencode('/home/sintu/Downloads/Project_pic_training/Bishwas_Tirkey/') # Folder where all your image files reside. Ensure it ends with '/
encoding_for_bishwas = [] # Create an empty list for saving encoded files
for b in os.listdir(my_dir1): # Loop over the folder to list individual files
    image1 = my_dir1 + b
    image1 = face_recognition.load_image_file(image1) # Run your load command
    image_encoding1 = face_recognition.face_encodings(image1) # Run your encoding command
    encoding_for_bishwas.append(image_encoding1[0]) # Append the results to encoding_for_file list
    
    
    
    
    
my_dir2 = os.fsencode('/home/sintu/Downloads/Project_pic_training/Ramkrishna/') # Folder where all your image files reside. Ensure it ends with '/
encoding_for_ramkrishna = [] # Create an empty list for saving encoded files
for c in os.listdir(my_dir2): # Loop over the folder to list individual files
    image2 = my_dir2 + c
    image2 = face_recognition.load_image_file(image2) # Run your load command
    image_encoding2 = face_recognition.face_encodings(image2) # Run your encoding command
    encoding_for_ramkrishna.append(image_encoding2[0]) # Append the results to encoding_for_file list    





my_dir3 = os.fsencode('/home/sintu/Downloads/Project_pic_training/Satyam_Kumar/') # Folder where all your image files reside. Ensure it ends with '/
encoding_for_satyam = [] # Create an empty list for saving encoded files
for d in os.listdir(my_dir3): # Loop over the folder to list individual files
    image3 = my_dir3 + d
    image3 = face_recognition.load_image_file(image3) # Run your load command
    image_encoding3 = face_recognition.face_encodings(image3) # Run your encoding command
    encoding_for_satyam.append(image_encoding3[0]) # Append the results to encoding_for_file list
    
    
    
    
my_dir4 = os.fsencode('/home/sintu/Downloads/Project_pic_training/Sintu_Kumar/') # Folder where all your image files reside. Ensure it ends with '/
encoding_for_sintu = [] # Create an empty list for saving encoded files
for e in os.listdir(my_dir4): # Loop over the folder to list individual files
    image4 = my_dir4 + e
    image4 = face_recognition.load_image_file(image4) # Run your load command
    image_encoding4 = face_recognition.face_encodings(image4) # Run your encoding command
    encoding_for_sintu.append(image_encoding4[0]) # Append the results to encoding_for_file list
  
    
    
my_dir5 = os.fsencode('/home/sintu/Downloads/Project_pic_training/Swadesh_Raj/') # Folder where all your image files reside. Ensure it ends with '/
encoding_for_swadesh = [] # Create an empty list for saving encoded files
for f in os.listdir(my_dir5): # Loop over the folder to list individual files
    image5 = my_dir5 + f
    image5 = face_recognition.load_image_file(image5) # Run your load command
    image_encoding5 = face_recognition.face_encodings(image5) # Run your encoding command
    encoding_for_swadesh.append(image_encoding5[0]) # Append the results to encoding_for_file list
    
 

#unknown_image = face_recognition.load_image_file('/home/swadesh/Downloads/Project/Test/test1.jpg')

#face_locations = face_recognition.face_locations(unknown_image,number_of_times_to_upsample=0, model="cnn")
#knownEncodings = []
#knownEncodings.append(encoding_for_ashutosh)
knownEncodings = [encoding_for_ashutosh, encoding_for_bishwas, encoding_for_ramkrishna,encoding_for_satyam, encoding_for_sintu, encoding_for_swadesh]
knownNames = ['Ashutosh Pandey','Bishwas Tirkey','Ramkrishna','Satyam Kumar','Sintu Kumar','Swadesh Raj']


data = {"encodings": knownEncodings, "names": knownNames}


image = cv2.imread('/home/sintu/Downloads/Project_pic_training/test/test1.jpg')
rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
 
# detect the (x, y)-coordinates of the bounding boxes corresponding
# to each face in the input image, then compute the facial embeddings
# for each face
#print("[INFO] recognizing faces...")
boxes = face_recognition.face_locations(rgb,
	model='hog')
encodings1 = face_recognition.face_encodings(rgb, boxes)
 
# initialize the list of names for each face detected
names = []

# loop over the facial embeddings
for encoding in encodings1:
    name = "Unknown"
    x=5
    for i in range(len(knownEncodings)):
        
        matches = face_recognition.compare_faces(knownEncodings[i],
                		encoding,tolerance=0.45)
        if True in matches:
             name = knownNames[i]
             names.append(name)
             x=0
             break
    if (x!=0):
        names.append(name)
print("Students present are as follows: ")   
for name in names:
     print(name)
            
#from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename = 'Attendance_List.xlsx')
ws = wb['Sheet1']
import pandas as pd
date = pd.datetime.now().date()
ws.cell(row = 1, column = 4, value = date)
for name in names:
    for i in range(2,ws.max_row):
        if(name == ws.cell(row=i, column=3).value):
            ws.cell(row=i, column=4).value = "Present"

wb.save("Attendance_List.xlsx")
